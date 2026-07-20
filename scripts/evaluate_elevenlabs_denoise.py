"""ElevenLabs ASR 去噪 vs 不去噪效果对比.

不降噪结果直接取 data/asr_testset/asr_test_pairs_elevenlabs.jsonl 的 asr 字段；
降噪结果用 DeepFilterNet 降噪后的音频重新调用 ElevenLabs speech-to-text。

Usage:
    python scripts/evaluate_elevenlabs_denoise.py
"""

from __future__ import annotations

import json
import os
import re
import sys
import time
from pathlib import Path
from typing import List, Optional

sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
import requests
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from src.audio_preprocessor import denoise_file
from tests.utils.metrics import cer

DATASET_ROOT = Path("dataset/语音识别样本集/录音普通话-长文本")
TESTSET = Path("data/asr_testset/asr_test_pairs_elevenlabs.jsonl")
DENOISED_DIR = Path("outputs/elevenlabs_denoised")
ELEVEN_BASE_URL = "https://api.elevenlabs.io"
ELEVEN_MODEL = os.getenv("ELEVENLABS_ASR_MODEL", "scribe_v1")


def load_api_key() -> str:
    key = os.getenv("ELEVENLABS_ASR_API_KEY", "")
    if key:
        return key
    # 从 .env.example 读取
    env_example = Path(".env.example")
    if env_example.exists():
        m = re.search(r"ELEVENLABS_ASR_API_KEY\s*=\s*(\S+)", env_example.read_text(encoding="utf-8"))
        if m:
            return m.group(1)
    return ""


def resolve_audio(record_id: str) -> Optional[Path]:
    parts = record_id.split("/")
    if len(parts) != 2:
        return None
    folder, name = parts
    candidate = DATASET_ROOT / folder / "录音" / f"{name}.wav"
    return candidate if candidate.exists() else None


def transcribe_elevenlabs(audio_path: Path, api_key: str, timeout: int = 120) -> str:
    url = f"{ELEVEN_BASE_URL}/v1/speech-to-text"
    headers = {"xi-api-key": api_key}
    with open(audio_path, "rb") as f:
        files = {"file": (audio_path.name, f, "audio/wav")}
        data = {
            "model_id": ELEVEN_MODEL,
            "language_code": "zh",
            "tag_audio_events": "false",
            "timestamps_granularity": "none",
        }
        response = requests.post(url, headers=headers, files=files, data=data, timeout=timeout)
    if response.status_code != 200:
        raise RuntimeError(f"ElevenLabs 请求失败: {response.status_code} {response.text[:500]}")
    return response.json().get("text", "")


def main():
    api_key = load_api_key()
    if not api_key:
        print("[ERROR] 未找到 ELEVENLABS_ASR_API_KEY")
        sys.exit(1)
    print(f"[INFO] ElevenLabs API key: {api_key[:8]}...")

    records = [json.loads(l) for l in TESTSET.read_text(encoding="utf-8").splitlines() if l.strip()]
    print(f"[INFO] 样本数: {len(records)}")

    DENOISED_DIR.mkdir(parents=True, exist_ok=True)
    rows = []
    stats = {"no_denoise": {"total": 0.0, "n": 0}, "denoise": {"total": 0.0, "n": 0}}

    for idx, r in enumerate(records, 1):
        rid = r["id"]
        correct = r["correct"]
        no_asr = r.get("asr", "")
        audio = resolve_audio(rid)
        if audio is None:
            print(f"  [{idx}/{len(records)}] 跳过 {rid}: 无音频")
            continue

        # 降噪
        denoised = DENOISED_DIR / f"{audio.stem}_att50.wav"
        if not denoised.exists():
            try:
                denoise_file(str(audio), str(denoised), method="deepfilternet")
            except Exception as e:
                print(f"  降噪失败 {rid}: {e}")
                denoised = audio

        print(f"  [{idx}/{len(records)}] {rid} 降噪转写中...")
        try:
            den_asr = transcribe_elevenlabs(denoised, api_key)
        except Exception as e:
            print(f"    ElevenLabs 降噪转写失败: {e}")
            den_asr = ""

        c_no = cer(no_asr, correct)
        c_den = cer(den_asr, correct)
        stats["no_denoise"]["total"] += c_no
        stats["no_denoise"]["n"] += 1
        stats["denoise"]["total"] += c_den
        stats["denoise"]["n"] += 1

        rows.append({
            "序号": idx,
            "ID": rid,
            "正确文本": correct,
            "不降噪转写": no_asr,
            "不降噪CER": round(c_no, 4),
            "降噪转写": den_asr,
            "降噪CER": round(c_den, 4),
            "CER变化": round(c_den - c_no, 4),
        })

    print(f"\n{'='*60}")
    print("ElevenLabs ASR 去噪 vs 不去噪（100 条）")
    print(f"{'='*60}")
    for mode in ["no_denoise", "denoise"]:
        s = stats[mode]
        avg = s["total"] / max(s["n"], 1)
        print(f"{mode:<14s} 平均CER: {avg:.4f}  样本数: {s['n']}")

    # 写 Excel
    ts = int(time.time())
    xlsx = Path("reports") / f"elevenlabs_denoise_compare_{ts}.xlsx"
    df = pd.DataFrame(rows)
    summary = pd.DataFrame([
        {"方式": "不降噪", "平均CER": round(stats["no_denoise"]["total"] / max(stats["no_denoise"]["n"], 1), 4), "样本数": stats["no_denoise"]["n"]},
        {"方式": "DeepFilterNet降噪", "平均CER": round(stats["denoise"]["total"] / max(stats["denoise"]["n"], 1), 4), "样本数": stats["denoise"]["n"]},
    ])
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="汇总", index=False)
        df.to_excel(writer, sheet_name="逐条对比", index=False)
        for ws in [writer.sheets["汇总"], writer.sheets["逐条对比"]]:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
        ws2 = writer.sheets["逐条对比"]
        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 16
        for col in ["C", "D", "E", "F", "G", "H"]:
            ws2.column_dimensions[col].width = 45
        for row_idx in range(2, len(df) + 2):
            for col_idx in [3, 4, 6]:
                ws2.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")
    print(f"[INFO] Excel 报告: {xlsx}")


if __name__ == "__main__":
    main()
