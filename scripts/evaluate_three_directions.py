"""全量验证LLM纠错效果.

使用ElevenLabs测试集（100条），对比四种模式：
1. baseline - 基线LLM纠错
2. nbest - 方向一：N-best候选融合
3. rag - 方向二：RAG增强
4. fusion - 融合模式：RAG + N-best

输出Excel报告（含条件格式高亮）：
- Sheet1 汇总指标
- Sheet2 逐条文本对比（原始/正确/规则/各模式LLM纠错结果+CER+状态）
"""

import argparse
import json
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.pipeline import PostCorrectionPipeline
from tests.utils.metrics import cer, is_valid_asr_text

MODES = ["baseline", "rag", "harness"]


def load_records(path: Path) -> list:
    records = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            if is_valid_asr_text(record.get("asr", "")):
                records.append(record)
    return records


def main():
    parser = argparse.ArgumentParser(description="评估LLM纠错效果")
    parser.add_argument(
        "--testset",
        type=str,
        default="data/asr_testset/asr_test_pairs_elevenlabs.jsonl",
        help="测试集路径",
    )
    args = parser.parse_args()

    testset = Path(args.testset)
    print(f"[INFO] 加载测试集: {testset}")
    records = load_records(testset)
    print(f"[INFO] 有效样本数: {len(records)}")

    print("[INFO] 初始化流水线...")
    pipeline = PostCorrectionPipeline()
    pipeline.warmup()

    n = len(records)

    # 汇总统计
    mode_stats = {m: {"total_cer_orig": 0.0, "total_cer_rule": 0.0, "total_cer_llm": 0.0,
                      "improved": 0, "degraded": 0, "unchanged": 0, "total_time": 0.0} for m in MODES}

    # 每条样本的详细结果（含完整文本）
    sample_details = []

    print(f"\n{'='*80}")
    print(f"开始评估（共{n}条样本，{len(MODES)}种模式）")
    print(f"{'='*80}")

    for idx, record in enumerate(records, 1):
        asr_text = record["asr"]
        correct_text = record["correct"]
        sample_id = record.get("id", "")

        # 规则纠错
        rule_result = pipeline.run(asr_text, layers=[1, 2, 3])
        rule_text = rule_result.corrected
        cer_orig = cer(asr_text, correct_text)
        cer_rule = cer(rule_text, correct_text)

        # 每条样本的详细记录
        detail = {
            "id": sample_id,
            "asr_original": asr_text,
            "correct": correct_text,
            "rule_corrected": rule_text,
            "cer_original": round(cer_orig, 4),
            "cer_rule": round(cer_rule, 4),
            "modes": {},
        }

        # 各LLM模式
        for mode in MODES:
            t0 = time.time()
            try:
                llm_result = pipeline.run(
                    asr_text, layers=[1, 2, 3],
                    enable_semantic=True, semantic_mode=mode,
                )
                llm_time = time.time() - t0
                llm_text = llm_result.corrected
                cer_llm = cer(llm_text, correct_text)
            except Exception as e:
                llm_time = time.time() - t0
                llm_text = rule_text
                cer_llm = cer_rule
                print(f"  [错误] {mode} idx={idx}: {type(e).__name__}: {e}")

            # 统计
            mode_stats[mode]["total_cer_orig"] += cer_orig
            mode_stats[mode]["total_cer_rule"] += cer_rule
            mode_stats[mode]["total_cer_llm"] += cer_llm
            mode_stats[mode]["total_time"] += llm_time

            if cer_llm < cer_rule - 0.001:
                mode_stats[mode]["improved"] += 1
                status = "改善"
            elif cer_llm > cer_rule + 0.001:
                mode_stats[mode]["degraded"] += 1
                status = "劣化"
            else:
                mode_stats[mode]["unchanged"] += 1
                status = "不变"

            detail["modes"][mode] = {
                "corrected": llm_text,
                "cer": round(cer_llm, 4),
                "status": status,
                "time_ms": round(llm_time * 1000, 1),
            }

        sample_details.append(detail)

        if idx % 50 == 0:
            print(f"  已处理 {idx}/{n}...")

    # 输出汇总报告
    print(f"\n{'='*100}")
    print("汇总报告")
    print(f"{'='*100}")
    print(f"{'模式':<12s} {'原始CER':>8s} {'规则CER':>8s} {'LLM CER':>8s} "
          f"{'LLM改善':>8s} {'改善':>6s} {'劣化':>6s} {'不变':>6s} {'平均耗时':>10s}")
    print("-" * 100)

    for mode in MODES:
        s = mode_stats[mode]
        avg_orig = s["total_cer_orig"] / n
        avg_rule = s["total_cer_rule"] / n
        avg_llm = s["total_cer_llm"] / n
        avg_time = s["total_time"] / n * 1000
        print(f"{mode:<12s} {avg_orig:>8.4f} {avg_rule:>8.4f} {avg_llm:>8.4f} "
              f"{avg_rule - avg_llm:>+8.4f} {s['improved']:>6d} {s['degraded']:>6d} "
              f"{s['unchanged']:>6d} {avg_time:>8.0f}ms")

    # 保存Excel报告
    reports_dir = Path("reports")
    reports_dir.mkdir(exist_ok=True)
    ts = int(time.time())

    # Sheet1: 汇总指标
    summary_rows = []
    for mode in MODES:
        s = mode_stats[mode]
        summary_rows.append({
            "模式": mode,
            "原始CER": round(s["total_cer_orig"] / n, 4),
            "规则CER": round(s["total_cer_rule"] / n, 4),
            "LLM CER": round(s["total_cer_llm"] / n, 4),
            "LLM改善": round(s["total_cer_rule"] / n - s["total_cer_llm"] / n, 4),
            "改善数": s["improved"],
            "劣化数": s["degraded"],
            "不变数": s["unchanged"],
            "平均耗时(ms)": round(s["total_time"] / n * 1000, 1),
        })
    df_summary = pd.DataFrame(summary_rows)

    # Sheet2: 逐条文本对比
    detail_rows = []
    for i, detail in enumerate(sample_details, 1):
        row = {
            "序号": i,
            "ID": detail["id"],
            "ASR原始": detail["asr_original"],
            "正确文本": detail["correct"],
            "规则纠错": detail["rule_corrected"],
            "原始CER": detail["cer_original"],
            "规则CER": detail["cer_rule"],
        }
        for mode in MODES:
            m = detail["modes"][mode]
            row[f"{mode}_纠错结果"] = m["corrected"]
            row[f"{mode}_CER"] = m["cer"]
            row[f"{mode}_状态"] = m["status"]
        detail_rows.append(row)
    df_detail = pd.DataFrame(detail_rows)

    # 写Excel
    xlsx_path = reports_dir / f"eval_report_{ts}.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="汇总指标", index=False)
        df_detail.to_excel(writer, sheet_name="逐条对比", index=False)

        # 获取workbook和worksheet做格式化
        wb = writer.book

        # === Sheet1 格式化 ===
        ws1 = writer.sheets["汇总指标"]
        ws1.column_dimensions["A"].width = 12
        for col_idx in range(2, 10):
            ws1.column_dimensions[get_column_letter(col_idx)].width = 12
        # 表头加粗
        for cell in ws1[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # === Sheet2 格式化 ===
        ws2 = writer.sheets["逐条对比"]
        # 列宽
        ws2.column_dimensions["A"].width = 6   # 序号
        ws2.column_dimensions["B"].width = 16   # ID
        ws2.column_dimensions["C"].width = 60   # ASR原始
        ws2.column_dimensions["D"].width = 60   # 正确文本
        ws2.column_dimensions["E"].width = 60   # 规则纠错
        ws2.column_dimensions["F"].width = 10   # 原始CER
        ws2.column_dimensions["G"].width = 10   # 规则CER

        col_offset = 8  # 从第H列开始是各模式数据
        for j, mode in enumerate(MODES):
            base = col_offset + j * 3
            ws2.column_dimensions[get_column_letter(base)].width = 60      # 纠错结果
            ws2.column_dimensions[get_column_letter(base + 1)].width = 10  # CER
            ws2.column_dimensions[get_column_letter(base + 2)].width = 8   # 状态

        # 表头加粗
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # 文本列自动换行
        text_cols = [3, 4, 5]  # C, D, E
        for j, mode in enumerate(MODES):
            text_cols.append(col_offset + j * 3)  # 各模式纠错结果列
        for row_idx in range(2, len(df_detail) + 2):
            for col_idx in text_cols:
                ws2.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

        # 条件格式：改善=绿色，劣化=红色
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        for row_idx in range(2, len(df_detail) + 2):
            for j, mode in enumerate(MODES):
                status_col = col_offset + j * 3 + 2  # 状态列
                cer_col = col_offset + j * 3 + 1     # CER列
                cell = ws2.cell(row=row_idx, column=status_col)
                cer_cell = ws2.cell(row=row_idx, column=cer_col)
                if cell.value == "改善":
                    cell.fill = green_fill
                    cer_cell.fill = green_fill
                elif cell.value == "劣化":
                    cell.fill = red_fill
                    cer_cell.fill = red_fill
                else:
                    cell.fill = yellow_fill

    print(f"[INFO] Excel报告已保存: {xlsx_path}")


if __name__ == "__main__":
    main()
