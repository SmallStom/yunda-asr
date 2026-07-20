"""qwen3-asr 提示/热词调优实验：在同一批降噪音频上对比不同 prompt 策略的原始 CER.

保留原始 transcribe_qwen3 逻辑不变，本脚本通过 prompt_mode 参数切换提示词，
用于选择最优 ASR 提示策略。

Usage:
    python scripts/evaluate_qwen3_prompt_tuning.py --testset data/asr_testset/asr_test_pairs_elevenlabs.jsonl
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path
from typing import List, Optional

sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.audio_preprocessor import denoise_file
from scripts.evaluate_qwen3_pipeline import load_hotwords
from tests.utils.metrics import cer


DATASET_ROOT = Path("dataset/语音识别样本集/录音普通话-长文本")
DEFAULT_BASE_URL = "http://192.168.1.119:8014"
DEFAULT_MODEL = "/models/Qwen3-ASR-1.7B"

PROMPT_MODES = ["original", "none", "domain", "domain_hotwords"]


def parse_args():
    parser = argparse.ArgumentParser(description="qwen3-asr 提示/热词调优")
    parser.add_argument("--testset", type=Path, default=Path("data/asr_testset/asr_test_pairs_elevenlabs.jsonl"))
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL)
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--output-dir", type=Path, default=Path("reports"))
    return parser.parse_args()


def load_records(path: Path, limit: int = 0) -> List[dict]:
    records = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                records.append(json.loads(line))
    if limit > 0:
        records = records[:limit]
    return records


def resolve_audio(record_id: str) -> Optional[Path]:
    parts = record_id.split("/")
    if len(parts) != 2:
        return None
    folder, name = parts
    candidate = DATASET_ROOT / folder / "录音" / f"{name}.wav"
    return candidate if candidate.exists() else None


def build_prompt(mode: str, hotwords: str) -> str:
    """根据模式生成 prompt 参数."""
    if mode == "none":
        return ""
    if mode == "original":
        if not hotwords.strip():
            return ""
        hw_list = [w.strip() for w in hotwords.replace("，", ",").split(",") if w.strip()][:100]
        return "请准确转写以下铁路调度相关音频。注意识别这些术语：" + "、".join(hw_list)
    if mode == "domain":
        return (
            "请准确转写铁路车站行车作业语音。要求：1. 铁路术语必须准确，"
            "如工务、电务、销记、道岔、联锁、进路、无表示、恢复表示、总锁闭、引导接车；"
            "2. 数字编号用阿拉伯数字，如13号道岔、3道、48615次、3G1；"
            "3. 保留书名号《行车设备检查登记簿》。"
        )
    if mode == "domain_hotwords":
        hw_list = [w.strip() for w in hotwords.replace("，", ",").split(",") if w.strip()][:100]
        return (
            "请准确转写铁路车站行车作业语音。要求：1. 铁路术语必须准确，"
            "如工务、电务、销记、道岔、联锁、进路、无表示、恢复表示、总锁闭、引导接车；"
            "2. 数字编号用阿拉伯数字，如13号道岔、3道、48615次、3G1；"
            "3. 保留书名号《行车设备检查登记簿》。"
            "重点注意这些术语：" + "、".join(hw_list)
        )
    return ""


def transcribe(audio_path: Path, prompt: str, base_url: str, model: str, timeout: int = 120) -> str:
    url = f"{base_url}/v1/audio/transcriptions"
    headers = {"Authorization": "Bearer dummy-key-for-local"}
    data = {
        "model": model,
        "language": "zh",
        "response_format": "json",
        "temperature": 0.0,
    }
    if prompt:
        data["prompt"] = prompt
    with open(audio_path, "rb") as f:
        files = {"file": (audio_path.name, f, "audio/wav")}
        response = requests.post(url, headers=headers, data=data, files=files, timeout=timeout)
    if response.status_code != 200:
        raise RuntimeError(f"qwen3-asr 请求失败: {response.status_code} {response.text[:800]}")
    return response.json().get("text", "")


def main():
    args = parse_args()
    hotwords = load_hotwords()
    print(f"[INFO] 加载热词 {len(hotwords.split(',')) if hotwords else 0} 个")

    records = load_records(args.testset, args.limit)
    print(f"[INFO] 测试样本数: {len(records)}")

    # 预生成每个样本的降噪音频，复用给所有 prompt 模式
    mode_stats = {m: {"total_cer": 0.0, "n": 0} for m in PROMPT_MODES}
    detail_rows = []

    n = len(records)
    for idx, record in enumerate(records, 1):
        rid = record.get("id", "")
        correct = record.get("correct", "")
        audio_path = resolve_audio(rid)
        if audio_path is None:
            print(f"  [{idx}/{n}] 跳过 {rid}: 找不到音频")
            continue

        denoised = Path("outputs/qwen3_prompt_tuning_denoised") / f"{audio_path.stem}_att50.wav"
        denoised.parent.mkdir(parents=True, exist_ok=True)
        if not denoised.exists():
            try:
                denoise_file(str(audio_path), str(denoised), method="deepfilternet")
            except Exception as e:
                print(f"  降噪失败 {rid}: {e}")
                denoised = audio_path

        print(f"  [{idx}/{n}] {rid}")
        row = {"序号": idx, "ID": rid, "正确文本": correct}
        for mode in PROMPT_MODES:
            prompt = build_prompt(mode, hotwords)
            try:
                text = transcribe(denoised, prompt, args.base_url, args.model)
            except Exception as e:
                print(f"    {mode} 失败: {e}")
                text = ""
            c = cer(text, correct)
            mode_stats[mode]["total_cer"] += c
            mode_stats[mode]["n"] += 1
            row[f"{mode}_转写"] = text
            row[f"{mode}_CER"] = round(c, 4)
        detail_rows.append(row)

    valid_n = mode_stats[PROMPT_MODES[0]]["n"]
    if valid_n == 0:
        print("[ERROR] 无成功样本")
        sys.exit(1)

    # 汇总
    print(f"\n{'='*80}")
    print("qwen3-asr 提示/热词调优汇总（降噪后）")
    print(f"{'='*80}")
    print(f"{'模式':<18s} {'平均CER':>8s}")
    print("-" * 30)
    for mode in PROMPT_MODES:
        s = mode_stats[mode]
        avg = s["total_cer"] / max(s["n"], 1)
        print(f"{mode:<18s} {avg:>8.4f}")

    # 写 Excel
    args.output_dir.mkdir(parents=True, exist_ok=True)
    ts = int(time.time())
    xlsx_path = args.output_dir / f"qwen3_prompt_tuning_{ts}.xlsx"

    summary_rows = [
        {"模式": m, "平均CER": round(mode_stats[m]["total_cer"] / max(mode_stats[m]["n"], 1), 4), "样本数": mode_stats[m]["n"]}
        for m in PROMPT_MODES
    ]
    df_summary = pd.DataFrame(summary_rows)
    df_detail = pd.DataFrame(detail_rows)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="汇总指标", index=False)
        df_detail.to_excel(writer, sheet_name="逐条对比", index=False)

        ws1 = writer.sheets["汇总指标"]
        ws1.column_dimensions["A"].width = 18
        ws1.column_dimensions["B"].width = 12
        ws1.column_dimensions["C"].width = 10
        for cell in ws1[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        ws2 = writer.sheets["逐条对比"]
        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 50
        col = 4
        for mode in PROMPT_MODES:
            ws2.column_dimensions[get_column_letter(col)].width = 50
            ws2.column_dimensions[get_column_letter(col + 1)].width = 10
            col += 2
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        text_cols = [3] + [4 + i * 2 for i in range(len(PROMPT_MODES))]
        for row_idx in range(2, len(df_detail) + 2):
            for col_idx in text_cols:
                ws2.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    print(f"[INFO] Excel 报告已保存: {xlsx_path}")


if __name__ == "__main__":
    main()
