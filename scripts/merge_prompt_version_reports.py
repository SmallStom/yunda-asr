"""合并v1和v2的eval_report，生成逐条对比报告.

用法:
    python scripts/merge_prompt_version_reports.py --v1 reports/eval_report_1782872434.xlsx --v2 reports/eval_report_1782872908.xlsx --output reports/prompt_v1_v2_compare.xlsx
"""

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def load_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet_name)


def status_tag(cer_v1: float, cer_v2: float) -> str:
    if cer_v2 < cer_v1 - 1e-9:
        return "v2改善"
    elif cer_v2 > cer_v1 + 1e-9:
        return "v2劣化"
    return "持平"


def main():
    parser = argparse.ArgumentParser(description="合并v1/v2评估报告")
    parser.add_argument("--v1", type=str, required=True, help="v1评估报告Excel路径")
    parser.add_argument("--v2", type=str, required=True, help="v2评估报告Excel路径")
    parser.add_argument("--output", type=str, default="reports/prompt_v1_v2_compare.xlsx", help="输出Excel路径")
    args = parser.parse_args()

    v1_path = Path(args.v1)
    v2_path = Path(args.v2)
    output = Path(args.output)

    # 读取逐条对比sheet
    df_v1 = load_sheet(v1_path, "逐条对比")
    df_v2 = load_sheet(v2_path, "逐条对比")

    # 读取汇总sheet
    sum_v1 = load_sheet(v1_path, "汇总指标")
    sum_v2 = load_sheet(v2_path, "汇总指标")

    # 合并
    merged = pd.DataFrame({
        "序号": df_v1["序号"],
        "ID": df_v1["ID"],
        "ASR原始": df_v1["ASR原始"],
        "正确文本": df_v1["正确文本"],
        "规则纠错": df_v1["规则纠错"],
        "规则CER": df_v1["规则CER"],
    })

    # 对每种模式，合并v1/v2
    modes = ["baseline", "rag", "harness"]
    for mode in modes:
        merged[f"{mode}_v1"] = df_v1[f"{mode}_纠错结果"]
        merged[f"{mode}_v1_CER"] = df_v1[f"{mode}_CER"]
        merged[f"{mode}_v2"] = df_v2[f"{mode}_纠错结果"]
        merged[f"{mode}_v2_CER"] = df_v2[f"{mode}_CER"]
        merged[f"{mode}差异"] = merged.apply(
            lambda row: status_tag(row[f"{mode}_v1_CER"], row[f"{mode}_v2_CER"]), axis=1
        )

    # 汇总对比
    summary_rows = []
    for mode in modes:
        row_v1 = sum_v1[sum_v1["模式"] == mode].iloc[0]
        row_v2 = sum_v2[sum_v2["模式"] == mode].iloc[0]
        summary_rows.append({
            "模式": mode,
            "v1_LLM_CER": row_v1["LLM CER"],
            "v2_LLM_CER": row_v2["LLM CER"],
            "CER变化": round(row_v2["LLM CER"] - row_v1["LLM CER"], 4),
            "v1_改善": row_v1["改善数"],
            "v2_改善": row_v2["改善数"],
            "v1_劣化": row_v1["劣化数"],
            "v2_劣化": row_v2["劣化数"],
            "v1_不变": row_v1["不变数"],
            "v2_不变": row_v2["不变数"],
            "v1_耗时": row_v1["平均耗时(ms)"],
            "v2_耗时": row_v2["平均耗时(ms)"],
        })
    df_summary = pd.DataFrame(summary_rows)

    # 写入Excel
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="汇总对比", index=False)
        merged.to_excel(writer, sheet_name="逐条对比", index=False)

        # 设置样式
        wb = writer.book
        ws_summary = wb["汇总对比"]
        ws_detail = wb["逐条对比"]

        # 汇总表头样式
        for cell in ws_summary[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 逐条对比表头样式
        for cell in ws_detail[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 条件格式：差异列
        diff_cols = [f"{mode}差异" for mode in modes]
        for col in ws_detail.iter_cols():
            col_letter = col[0].column_letter
            col_name = ws_detail.cell(row=1, column=col[0].column).value
            if col_name in diff_cols:
                for cell in col[1:]:
                    if cell.value == "v2改善":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100")
                    elif cell.value == "v2劣化":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")

        # 自动列宽
        for ws in [ws_summary, ws_detail]:
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        cell_len = len(str(cell.value))
                        if cell_len > max_length:
                            max_length = cell_len
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                ws.column_dimensions[col_letter].width = adjusted_width

        # 文本列自动换行
        for row in ws_detail.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and len(cell.value) > 10:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

    print(f"[INFO] 对比报告已保存: {output}")


if __name__ == "__main__":
    main()
