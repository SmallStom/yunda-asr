"""qwen3-asr 热词调用方式实验：通过 /v1/chat/completions 的 system 消息注入热词/领域上下文.

之前用 /v1/audio/transcriptions 的 prompt 参数无效果，本脚本验证官方推荐的方式：
将领域上下文/热词放在 system 消息中，音频放在 user 消息中。

Usage:
    python scripts/evaluate_qwen3_hotwords_chat.py --limit 5
"""

from __future__ import annotations

import argparse
import base64
import json
import re
import sys
import time
from pathlib import Path
from typing import List, Optional

sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
import requests
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from scripts.evaluate_qwen3_pipeline import load_hotwords
from tests.utils.metrics import cer

DATASET_ROOT = Path("dataset/语音识别样本集/录音普通话-长文本")
DEFAULT_BASE_URL = "http://192.168.1.119:8014"
DEFAULT_MODEL = "/models/Qwen3-ASR-1.7B"

PROMPT_MODES = ["none", "hotwords_20_comma", "context_paragraph", "hotwords_100_comma"]

# 20 个核心铁路术语，避免过长上下文干扰
CORE_HOTWORDS = [
    "工务", "电务", "销记", "道岔", "联锁", "进路", "无表示", "恢复表示",
    "总锁闭", "引导接车", "手摇把", "转辙机", "占线簿", "加岗", "登记簿",
    "轨道电路", "红光带", "扳动", "密贴", "凭证",
]


def parse_args():
    parser = argparse.ArgumentParser(description="qwen3-asr 热词 chat 方式实验")
    parser.add_argument("--testset", type=Path, default=Path("data/asr_testset/asr_test_pairs_elevenlabs.jsonl"))
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL)
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--output-dir", type=Path, default=Path("reports"))
    return parser.parse_args()


def load_records(path: Path, limit: int = 0) -> List[dict]:
    records = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                records.append(json.loads(line))
    if limit > 0:
        records = records[:limit]
    return records


def resolve_audio(record_id: str) -> Optional[Path]:
    parts = record_id.split("/")
    if len(parts) != 2:
        return None
    folder, name = parts
    candidate = DATASET_ROOT / folder / "录音" / f"{name}.wav"
    return candidate if candidate.exists() else None


def build_system(mode: str, hotwords: str) -> str:
    if mode == "none":
        return ""
    if mode == "hotwords_20_comma":
        return ",".join(CORE_HOTWORDS)
    if mode == "context_paragraph":
        return "铁路车站行车作业语音，涉及" + "、".join(CORE_HOTWORDS) + "等术语。"
    if mode == "hotwords_100_comma":
        hw_list = [w.strip() for w in hotwords.replace("，", ",").split(",") if w.strip()][:100]
        return ",".join(hw_list)
    return ""


def extract_text(content: str) -> str:
    content = content.strip()
    content = re.sub(r"^language\s+[A-Za-z]+", "", content)
    content = content.replace("<asr_text>", "").replace("</asr_text>", "")
    return content.strip()


def transcribe_chat(audio_path: Path, system_prompt: str, base_url: str, model: str, timeout: int = 120) -> str:
    url = f"{base_url}/v1/chat/completions"
    with open(audio_path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("ascii")
    data_url = f"data:audio/wav;base64,{b64}"
    messages = []
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    messages.append({
        "role": "user",
        "content": [{"type": "audio_url", "audio_url": {"url": data_url}}]
    })
    payload = {
        "model": model,
        "messages": messages,
        "temperature": 0.0,
        "max_tokens": 2048,
        "chat_template_kwargs": {"enable_thinking": False},
    }
    headers = {"Content-Type": "application/json", "Authorization": "Bearer dummy-key-for-local"}
    response = requests.post(url, json=payload, headers=headers, timeout=timeout)
    if response.status_code != 200:
        raise RuntimeError(f"chat 请求失败: {response.status_code} {response.text[:800]}")
    content = response.json()["choices"][0]["message"]["content"]
    return extract_text(content)


def main():
    args = parse_args()
    hotwords = load_hotwords()
    print(f"[INFO] 热词数: {len(hotwords.split(',')) if hotwords else 0}")

    records = load_records(args.testset, args.limit)
    print(f"[INFO] 样本数: {len(records)}")

    mode_stats = {m: {"total_cer": 0.0, "n": 0} for m in PROMPT_MODES}
    detail_rows = []

    for idx, record in enumerate(records, 1):
        rid = record.get("id", "")
        correct = record.get("correct", "")
        audio_path = resolve_audio(rid)
        if audio_path is None:
            continue

        # 复用已有降噪文件
        denoised = Path("outputs/qwen3_prompt_tuning_denoised") / f"{audio_path.stem}_att50.wav"
        if not denoised.exists():
            denoised = Path("outputs/qwen3_hotwords_denoised") / f"{audio_path.stem}_att50.wav"
            denoised.parent.mkdir(parents=True, exist_ok=True)
            from src.audio_preprocessor import denoise_file
            try:
                denoise_file(str(audio_path), str(denoised), method="deepfilternet")
            except Exception as e:
                print(f"  降噪失败 {rid}: {e}")
                denoised = audio_path

        print(f"  [{idx}/{len(records)}] {rid}")
        row = {"序号": idx, "ID": rid, "正确文本": correct}
        for mode in PROMPT_MODES:
            system_prompt = build_system(mode, hotwords)
            try:
                text = transcribe_chat(denoised, system_prompt, args.base_url, args.model)
            except Exception as e:
                print(f"    {mode} 失败: {e}")
                text = ""
            c = cer(text, correct)
            mode_stats[mode]["total_cer"] += c
            mode_stats[mode]["n"] += 1
            row[f"{mode}_转写"] = text
            row[f"{mode}_CER"] = round(c, 4)
        detail_rows.append(row)

    valid_n = mode_stats[PROMPT_MODES[0]]["n"]
    if valid_n == 0:
        print("[ERROR] 无成功样本")
        sys.exit(1)

    print(f"\n{'='*70}")
    print("qwen3-asr 热词 chat 方式汇总（降噪后）")
    print(f"{'='*70}")
    print(f"{'模式':<18s} {'平均CER':>8s}")
    print("-" * 30)
    for mode in PROMPT_MODES:
        s = mode_stats[mode]
        avg = s["total_cer"] / max(s["n"], 1)
        print(f"{mode:<18s} {avg:>8.4f}")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    ts = int(time.time())
    xlsx_path = args.output_dir / f"qwen3_hotwords_chat_{ts}.xlsx"

    summary_rows = [
        {"模式": m, "平均CER": round(mode_stats[m]["total_cer"] / max(mode_stats[m]["n"], 1), 4), "样本数": mode_stats[m]["n"]}
        for m in PROMPT_MODES
    ]
    df_summary = pd.DataFrame(summary_rows)
    df_detail = pd.DataFrame(detail_rows)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="汇总指标", index=False)
        df_detail.to_excel(writer, sheet_name="逐条对比", index=False)
        for ws in [writer.sheets["汇总指标"], writer.sheets["逐条对比"]]:
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
        ws2 = writer.sheets["逐条对比"]
        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 50
        col = 4
        for _ in PROMPT_MODES:
            ws2.column_dimensions[get_column_letter(col)].width = 50
            ws2.column_dimensions[get_column_letter(col + 1)].width = 10
            col += 2

    print(f"[INFO] Excel 报告: {xlsx_path}")


if __name__ == "__main__":
    main()
