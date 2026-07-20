"""qwen3-asr (vLLM) + DeepFilterNet 降噪 + 后处理纠错 全链路评估.

Usage:
    python scripts/evaluate_qwen3_pipeline.py --limit 3
    python scripts/evaluate_qwen3_pipeline.py --testset data/asr_testset/asr_test_pairs_long.jsonl
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path
from typing import List, Optional

sys.path.insert(0, str(Path(__file__).parent.parent))

import requests

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.audio_preprocessor import denoise_file
from src.pipeline import PostCorrectionPipeline
from tests.utils.metrics import cer, entity_fidelity, term_accuracy


DATASET_ROOT = Path("dataset/语音识别样本集/录音普通话-长文本")
HOTWORDS_PATH = Path("data/lexicon/hotwords.json")
DEFAULT_BASE_URL = "http://192.168.1.119:8014"
DEFAULT_MODEL = "/models/Qwen3-ASR-1.7B"


def load_hotwords(path: Path = HOTWORDS_PATH, max_words: int = 300) -> str:
    """加载热词文件，返回逗号分隔字符串."""
    if not path.exists():
        return ""
    if path.suffix == ".json":
        words = json.loads(path.read_text(encoding="utf-8"))
    else:
        words = [w.strip() for w in path.read_text(encoding="utf-8").splitlines() if w.strip()]
    return ",".join(words[:max_words])


def load_records(path: Path, limit: int = 0, subdirs: Optional[List[str]] = None) -> List[dict]:
    records = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            if subdirs:
                rid = record.get("id", "")
                if not any(rid.startswith(d + "/") for d in subdirs):
                    continue
            records.append(record)
    if limit > 0:
        records = records[:limit]
    return records


def resolve_audio(record_id: str) -> Optional[Path]:
    """由 id 如 '1/train_0008' 解析音频路径."""
    parts = record_id.split("/")
    if len(parts) != 2:
        return None
    folder, name = parts
    candidate = DATASET_ROOT / folder / "录音" / f"{name}.wav"
    return candidate if candidate.exists() else None


def transcribe_qwen3(audio_path: Path, hotwords: str, base_url: str, model: str, timeout: int = 120) -> str:
    """调用 qwen3-asr (vLLM /v1/audio/transcriptions)."""
    url = f"{base_url}/v1/audio/transcriptions"
    headers = {"Authorization": "Bearer dummy-key-for-local"}

    data = {
        "model": model,
        "language": "zh",
        "response_format": "json",
        "temperature": 0.0,
    }
    if hotwords.strip():
        hw_list = [w.strip() for w in hotwords.replace("，", ",").split(",") if w.strip()][:100]
        data["prompt"] = "请准确转写以下铁路调度相关音频。注意识别这些术语：" + "、".join(hw_list)

    with open(audio_path, "rb") as f:
        files = {"file": (audio_path.name, f, "audio/wav")}
        response = requests.post(url, headers=headers, data=data, files=files, timeout=timeout)

    if response.status_code != 200:
        raise RuntimeError(f"qwen3-asr 请求失败: {response.status_code} {response.text[:800]}")

    return response.json().get("text", "")


def evaluate(
    records: List[dict],
    pipeline: PostCorrectionPipeline,
    hotwords: str,
    base_url: str,
    model: str,
    denoise: bool,
    enable_semantic: bool,
) -> dict:
    results = []
    total_cer_raw = 0.0
    total_cer_corrected = 0.0
    total_term_raw = 0.0
    total_term_corrected = 0.0
    total_entity_raw = 0.0
    total_entity_corrected = 0.0
    total_asr_latency = 0.0
    total_pipeline_latency = 0.0
    failed = 0

    layer_stats = {
        "preprocessor": {"triggered": 0, "samples": 0},
        "dictionary": {"triggered": 0, "samples": 0},
        "context": {"triggered": 0, "samples": 0},
        "semantic": {"triggered": 0, "samples": 0},
    }

    n = len(records)
    for idx, record in enumerate(records, 1):
        rid = record.get("id", "")
        correct = record.get("correct", "")
        audio_path = resolve_audio(rid)
        if audio_path is None:
            print(f"  [{idx}/{n}] 跳过 {rid}: 找不到音频")
            failed += 1
            continue

        print(f"  [{idx}/{n}] {rid} -> {audio_path.name}")

        # 1. 可选 DeepFilterNet 降噪
        input_audio = audio_path
        if denoise:
            denoised_path = Path("outputs/qwen3_eval_denoised") / f"{audio_path.stem}_att50.wav"
            denoised_path.parent.mkdir(parents=True, exist_ok=True)
            try:
                denoise_file(str(audio_path), str(denoised_path), method="deepfilternet")
                input_audio = denoised_path
            except Exception as e:
                print(f"    降噪失败，使用原音: {e}")

        # 2. ASR
        t0 = time.time()
        try:
            raw_text = transcribe_qwen3(input_audio, hotwords, base_url, model)
        except Exception as e:
            print(f"    ASR失败: {e}")
            failed += 1
            continue
        asr_latency = time.time() - t0
        total_asr_latency += asr_latency

        # 3. 后处理纠错
        t1 = time.time()
        result = pipeline.run(
            raw_text,
            layers=[1, 2, 3],
            enable_semantic=enable_semantic,
            semantic_mode="rag",
        )
        pipeline_latency = time.time() - t1
        total_pipeline_latency += pipeline_latency
        corrected = result.corrected

        # 4. 指标
        cer_raw = cer(raw_text, correct)
        cer_corrected = cer(corrected, correct)
        term_raw, _, _ = term_accuracy(raw_text, correct)
        term_corrected, _, _ = term_accuracy(corrected, correct)
        entity_raw, _, _ = entity_fidelity(raw_text, correct)
        entity_corrected, _, _ = entity_fidelity(corrected, correct)

        total_cer_raw += cer_raw
        total_cer_corrected += cer_corrected
        total_term_raw += term_raw
        total_term_corrected += term_corrected
        total_entity_raw += entity_raw
        total_entity_corrected += entity_corrected

        asr_latency_ms = round(asr_latency * 1000, 1)
        pipeline_latency_ms = round(pipeline_latency * 1000, 1)

        for layer in result.layers_applied:
            if layer in layer_stats:
                layer_stats[layer]["triggered"] += 1
        for layer in layer_stats:
            layer_stats[layer]["samples"] += 1

        results.append({
            "id": rid,
            "audio": str(audio_path),
            "correct": correct,
            "asr_raw": raw_text,
            "asr_corrected": corrected,
            "cer_raw": round(cer_raw, 4),
            "cer_corrected": round(cer_corrected, 4),
            "layers_applied": result.layers_applied,
            "asr_latency_ms": asr_latency_ms,
            "pipeline_latency_ms": pipeline_latency_ms,
        })

    valid_n = len(results)
    report = {
        "meta": {
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "asr_service": "qwen3-asr (vLLM)",
            "base_url": base_url,
            "model": model,
            "denoise": denoise,
            "enable_semantic": enable_semantic,
            "total_samples": n,
            "successful": valid_n,
            "failed": failed,
        },
        "summary": {
            "avg_cer_raw": round(total_cer_raw / valid_n, 4) if valid_n else 0,
            "avg_cer_corrected": round(total_cer_corrected / valid_n, 4) if valid_n else 0,
            "avg_term_accuracy_raw": round(total_term_raw / valid_n, 4) if valid_n else 0,
            "avg_term_accuracy_corrected": round(total_term_corrected / valid_n, 4) if valid_n else 0,
            "avg_entity_fidelity_raw": round(total_entity_raw / valid_n, 4) if valid_n else 0,
            "avg_entity_fidelity_corrected": round(total_entity_corrected / valid_n, 4) if valid_n else 0,
            "avg_asr_latency_ms": round(total_asr_latency / valid_n * 1000, 1) if valid_n else 0,
            "avg_pipeline_latency_ms": round(total_pipeline_latency / valid_n * 1000, 1) if valid_n else 0,
        },
        "layer_stats": {
            layer: {
                "triggered": stats["triggered"],
                "trigger_rate": round(stats["triggered"] / max(stats["samples"], 1), 4),
            }
            for layer, stats in layer_stats.items()
        },
        "details": results,
    }
    return report


def export_excel_report(report: dict, excel_path: Path) -> None:
    """把 qwen3-asr 全链路评估结果导出为 Excel 对比表."""
    meta = report["meta"]
    summary = report["summary"]

    # Sheet1: 汇总指标
    summary_rows = [
        {
            "总样本数": meta["total_samples"],
            "成功数": meta["successful"],
            "失败数": meta["failed"],
            "ASR服务": meta["asr_service"],
            "模型": meta["model"],
            "降噪": meta["denoise"],
            "语义精修": meta["enable_semantic"],
            "CER_ASR原始": summary["avg_cer_raw"],
            "CER_纠错后": summary["avg_cer_corrected"],
            "CER_改善": round(summary["avg_cer_raw"] - summary["avg_cer_corrected"], 4),
            "术语准确率_ASR原始": summary["avg_term_accuracy_raw"],
            "术语准确率_纠错后": summary["avg_term_accuracy_corrected"],
            "实体保真率_ASR原始": summary["avg_entity_fidelity_raw"],
            "实体保真率_纠错后": summary["avg_entity_fidelity_corrected"],
            "平均ASR延迟_ms": summary["avg_asr_latency_ms"],
            "平均纠错延迟_ms": summary["avg_pipeline_latency_ms"],
        }
    ]
    df_summary = pd.DataFrame(summary_rows)

    # Sheet2: 各层触发
    layer_rows = [
        {
            "层": layer,
            "触发次数": stats["triggered"],
            "触发率": stats["trigger_rate"],
        }
        for layer, stats in report.get("layer_stats", {}).items()
    ]
    df_layers = pd.DataFrame(layer_rows)

    # Sheet3: 逐条对比
    detail_rows = []
    for r in report["details"]:
        cer_raw = r["cer_raw"]
        cer_corrected = r["cer_corrected"]
        if cer_corrected < cer_raw - 0.001:
            status = "改善"
        elif cer_corrected > cer_raw + 0.001:
            status = "劣化"
        else:
            status = "不变"
        detail_rows.append(
            {
                "序号": len(detail_rows) + 1,
                "ID": r["id"],
                "音频路径": r["audio"],
                "正确文本": r["correct"],
                "qwen3-asr原始": r["asr_raw"],
                "纠错后": r["asr_corrected"],
                "CER_ASR原始": cer_raw,
                "CER_纠错后": cer_corrected,
                "状态": status,
                "应用层": ",".join(str(x) for x in r.get("layers_applied", [])),
                "ASR延迟_ms": r.get("asr_latency_ms", 0),
                "纠错延迟_ms": r.get("pipeline_latency_ms", 0),
            }
        )
    df_detail = pd.DataFrame(detail_rows)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="汇总指标", index=False)
        df_layers.to_excel(writer, sheet_name="各层触发", index=False)
        df_detail.to_excel(writer, sheet_name="逐条对比", index=False)

        wb = writer.book

        # 表头样式
        header_font = Font(bold=True)
        header_align = Alignment(horizontal="center", wrap_text=True)

        # Sheet1 格式
        ws1 = writer.sheets["汇总指标"]
        for col_idx in range(1, df_summary.shape[1] + 1):
            ws1.column_dimensions[get_column_letter(col_idx)].width = 16
        for cell in ws1[1]:
            cell.font = header_font
            cell.alignment = header_align

        # Sheet2 格式
        ws_layers = writer.sheets["各层触发"]
        ws_layers.column_dimensions["A"].width = 16
        ws_layers.column_dimensions["B"].width = 12
        ws_layers.column_dimensions["C"].width = 12
        for cell in ws_layers[1]:
            cell.font = header_font
            cell.alignment = header_align

        # Sheet3 格式
        ws3 = writer.sheets["逐条对比"]
        ws3.column_dimensions["A"].width = 8   # 序号
        ws3.column_dimensions["B"].width = 16  # ID
        ws3.column_dimensions["C"].width = 45  # 音频路径
        ws3.column_dimensions["D"].width = 50  # 正确文本
        ws3.column_dimensions["E"].width = 50  # qwen3-asr原始
        ws3.column_dimensions["F"].width = 50  # 纠错后
        ws3.column_dimensions["G"].width = 12  # CER_ASR原始
        ws3.column_dimensions["H"].width = 12  # CER_纠错后
        ws3.column_dimensions["I"].width = 10  # 状态
        ws3.column_dimensions["J"].width = 16  # 应用层
        ws3.column_dimensions["K"].width = 12  # ASR延迟
        ws3.column_dimensions["L"].width = 12  # 纠错延迟

        for cell in ws3[1]:
            cell.font = header_font
            cell.alignment = header_align

        text_cols = [3, 4, 5, 6]  # C, D, E, F
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        for row_idx in range(2, len(df_detail) + 2):
            for col_idx in text_cols:
                ws3.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )
            status_cell = ws3.cell(row=row_idx, column=9)
            cer_after_cell = ws3.cell(row=row_idx, column=8)
            if status_cell.value == "改善":
                status_cell.fill = green_fill
                cer_after_cell.fill = green_fill
            elif status_cell.value == "劣化":
                status_cell.fill = red_fill
                cer_after_cell.fill = red_fill
            else:
                status_cell.fill = yellow_fill
                cer_after_cell.fill = yellow_fill

    print(f"[INFO] Excel 报告已保存: {excel_path}")


def main():
    parser = argparse.ArgumentParser(description="qwen3-asr 全链路评估")
    parser.add_argument("--testset", type=Path, default=Path("data/asr_testset/asr_test_pairs_long.jsonl"))
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL)
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument(
        "--subdirs",
        nargs="+",
        type=str,
        default=None,
        help="按子目录过滤 (如: 1 2 3)",
    )
    parser.add_argument("--no-denoise", action="store_true", help="禁用 DeepFilterNet 降噪")
    parser.add_argument("--enable-semantic", action="store_true", help="启用 Layer 4 语义精修")
    parser.add_argument("--output-dir", type=Path, default=Path("reports"))
    args = parser.parse_args()

    hotwords = load_hotwords()
    print(f"[INFO] 加载热词 {len(hotwords.split(',')) if hotwords else 0} 个")

    records = load_records(args.testset, args.limit, subdirs=args.subdirs)
    print(f"[INFO] 测试样本数: {len(records)}")
    if not records:
        print("[ERROR] 无测试样本")
        sys.exit(1)

    print("[INFO] 初始化后处理流水线...")
    pipeline = PostCorrectionPipeline()
    pipeline.warmup()

    print(f"[INFO] 开始评估 (denoise={not args.no_denoise}, semantic={args.enable_semantic})...")
    t0 = time.time()
    report = evaluate(
        records,
        pipeline,
        hotwords,
        args.base_url,
        args.model,
        denoise=not args.no_denoise,
        enable_semantic=args.enable_semantic,
    )
    elapsed = time.time() - t0

    summary = report["summary"]
    print("\n" + "=" * 50)
    print("qwen3-asr 全链路评估摘要")
    print("=" * 50)
    print(f"总样本: {report['meta']['total_samples']} | 成功: {report['meta']['successful']} | 失败: {report['meta']['failed']}")
    print(f"平均 CER:  {summary['avg_cer_raw']:.4f} -> {summary['avg_cer_corrected']:.4f}")
    print(f"术语准确率: {summary['avg_term_accuracy_raw']:.4f} -> {summary['avg_term_accuracy_corrected']:.4f}")
    print(f"实体保真率: {summary['avg_entity_fidelity_raw']:.4f} -> {summary['avg_entity_fidelity_corrected']:.4f}")
    print(f"平均延迟: ASR {summary['avg_asr_latency_ms']:.0f}ms | 纠错 {summary['avg_pipeline_latency_ms']:.0f}ms")
    print(f"总耗时: {elapsed:.1f}s")
    print("=" * 50)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    report_path = args.output_dir / f"qwen3_pipeline_{timestamp}.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"[INFO] 报告已保存: {report_path}")

    excel_path = args.output_dir / f"qwen3_pipeline_{timestamp}.xlsx"
    export_excel_report(report, excel_path)


if __name__ == "__main__":
    main()
