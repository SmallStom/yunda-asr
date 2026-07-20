"""基于真实ASR样本集的批量评估脚本.

Usage:
    python scripts/evaluate_on_testset.py --layers 1 2 3 --output-dir reports/
    python scripts/evaluate_on_testset.py --enable-semantic --subdirs 1 2
"""

import argparse
import json
import sys
import time
from pathlib import Path
from typing import List, Optional

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.pipeline import PostCorrectionPipeline
from tests.utils.metrics import cer, entity_fidelity, is_valid_asr_text, term_accuracy

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def parse_args():
    parser = argparse.ArgumentParser(description="ASR纠错系统批量评估")
    parser.add_argument(
        "--testset",
        type=Path,
        default=Path("data/asr_testset/asr_test_pairs.jsonl"),
        help="测试集JSONL文件路径",
    )
    parser.add_argument(
        "--layers",
        nargs="+",
        type=int,
        default=[1, 2, 3],
        help="启用的层号列表 (默认: 1 2 3)",
    )
    parser.add_argument(
        "--enable-semantic",
        action="store_true",
        help="是否启用Layer 4语义精修",
    )
    parser.add_argument(
        "--subdirs",
        nargs="+",
        type=str,
        default=None,
        help="按子目录过滤 (如: 1 2 3 4)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="限制处理样本数 (0=不限制)",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("reports"),
        help="报告输出目录",
    )
    return parser.parse_args()


def load_records(path: Path, subdirs: Optional[List[str]] = None) -> List[dict]:
    """加载测试记录，过滤无效样本."""
    records = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            asr_text = record.get("asr", "")
            if not is_valid_asr_text(asr_text):
                continue
            if subdirs:
                rid = record.get("id", "")
                if not any(rid.startswith(d + "/") for d in subdirs):
                    continue
            records.append(record)
    return records


def evaluate(
    records: List[dict],
    pipeline: PostCorrectionPipeline,
    layers: List[int],
    enable_semantic: bool,
) -> dict:
    """执行评估，返回报告字典."""
    results = []
    total_cer_before = 0.0
    total_cer_after = 0.0
    total_term_acc_before = 0.0
    total_term_acc_after = 0.0
    total_entity_fid_before = 0.0
    total_entity_fid_after = 0.0
    improved_count = 0
    unchanged_count = 0
    degraded_count = 0
    total_latency = 0.0

    layer_stats = {
        "preprocessor": {"triggered": 0, "samples": 0},
        "dictionary": {"triggered": 0, "samples": 0},
        "context": {"triggered": 0, "samples": 0},
        "semantic": {"triggered": 0, "samples": 0},
    }

    n = len(records)
    for idx, record in enumerate(records, 1):
        asr_text = record["asr"]
        correct_text = record["correct"]

        t0 = time.time()
        result = pipeline.run(
            asr_text,
            layers=layers,
            enable_semantic=enable_semantic,
        )
        latency = time.time() - t0
        total_latency += latency

        cer_before = cer(asr_text, correct_text)
        cer_after = cer(result.corrected, correct_text)

        term_acc_before, _, _ = term_accuracy(asr_text, correct_text)
        term_acc_after, term_hits, term_total = term_accuracy(result.corrected, correct_text)

        entity_fid_before, _, _ = entity_fidelity(asr_text, correct_text)
        entity_fid_after, entity_hits, entity_total = entity_fidelity(result.corrected, correct_text)

        total_cer_before += cer_before
        total_cer_after += cer_after
        total_term_acc_before += term_acc_before
        total_term_acc_after += term_acc_after
        total_entity_fid_before += entity_fid_before
        total_entity_fid_after += entity_fid_after

        if cer_after < cer_before - 0.001:
            improved_count += 1
        elif cer_after > cer_before + 0.001:
            degraded_count += 1
        else:
            unchanged_count += 1

        for layer in result.layers_applied:
            if layer in layer_stats:
                layer_stats[layer]["triggered"] += 1
        for layer in layer_stats:
            layer_stats[layer]["samples"] += 1

        results.append({
            "id": record.get("id", ""),
            "asr": asr_text,
            "correct": correct_text,
            "corrected": result.corrected,
            "cer_before": round(cer_before, 4),
            "cer_after": round(cer_after, 4),
            "term_hits": term_hits,
            "term_total": term_total,
            "entity_hits": entity_hits,
            "entity_total": entity_total,
            "layers_applied": result.layers_applied,
            "latency_ms": round(latency * 1000, 2),
        })

        if idx % 50 == 0:
            print(f"  已处理 {idx}/{n} 条...")

    avg_cer_before = total_cer_before / n if n else 0
    avg_cer_after = total_cer_after / n if n else 0
    avg_latency = total_latency / n if n else 0

    report = {
        "meta": {
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "total_samples": n,
            "layers": layers,
            "enable_semantic": enable_semantic,
        },
        "summary": {
            "avg_cer_before": round(avg_cer_before, 4),
            "avg_cer_after": round(avg_cer_after, 4),
            "cer_improvement": round(avg_cer_before - avg_cer_after, 4),
            "avg_term_accuracy_before": round(total_term_acc_before / n if n else 0, 4),
            "avg_term_accuracy_after": round(total_term_acc_after / n if n else 0, 4),
            "avg_entity_fidelity_before": round(total_entity_fid_before / n if n else 0, 4),
            "avg_entity_fidelity_after": round(total_entity_fid_after / n if n else 0, 4),
            "improved_count": improved_count,
            "unchanged_count": unchanged_count,
            "degraded_count": degraded_count,
            "avg_latency_ms": round(avg_latency * 1000, 2),
        },
        "layer_stats": {
            layer: {
                "triggered": stats["triggered"],
                "trigger_rate": round(stats["triggered"] / max(stats["samples"], 1), 4),
            }
            for layer, stats in layer_stats.items()
        },
        "worst_cases": [
            {
                "id": r["id"],
                "asr": r["asr"],
                "correct": r["correct"],
                "corrected": r["corrected"],
                "cer_before": r["cer_before"],
                "cer_after": r["cer_after"],
            }
            for r in sorted(
                [r for r in results if r["cer_after"] > r["cer_before"]],
                key=lambda x: x["cer_after"] - x["cer_before"],
                reverse=True,
            )[:10]
        ],
        "best_cases": [
            {
                "id": r["id"],
                "asr": r["asr"],
                "correct": r["correct"],
                "corrected": r["corrected"],
                "cer_before": r["cer_before"],
                "cer_after": r["cer_after"],
            }
            for r in sorted(
                [r for r in results if r["cer_after"] < r["cer_before"]],
                key=lambda x: x["cer_before"] - x["cer_after"],
                reverse=True,
            )[:10]
        ],
        "details": results,
    }

    return report


def export_excel_report(report: dict, excel_path: Path) -> None:
    """把评估报告导出为带条件格式的 Excel."""
    summary = report["summary"]
    meta = report["meta"]

    # Sheet1: 汇总指标
    summary_rows = [
        {
            "总样本数": meta["total_samples"],
            "启用层": ",".join(str(x) for x in meta["layers"]),
            "语义精修": meta["enable_semantic"],
            "CER_纠错前": summary["avg_cer_before"],
            "CER_纠错后": summary["avg_cer_after"],
            "CER_改善": summary["cer_improvement"],
            "术语准确率_前": summary["avg_term_accuracy_before"],
            "术语准确率_后": summary["avg_term_accuracy_after"],
            "实体保真率_前": summary["avg_entity_fidelity_before"],
            "实体保真率_后": summary["avg_entity_fidelity_after"],
            "改善数": summary["improved_count"],
            "不变数": summary["unchanged_count"],
            "劣化数": summary["degraded_count"],
            "平均延迟_ms": summary["avg_latency_ms"],
        }
    ]
    df_summary = pd.DataFrame(summary_rows)

    layer_rows = [
        {
            "层": layer,
            "触发次数": stats["triggered"],
            "触发率": stats["trigger_rate"],
        }
        for layer, stats in report.get("layer_stats", {}).items()
    ]
    df_layers = pd.DataFrame(layer_rows)

    # Sheet2: 逐条对比
    detail_rows = []
    for r in report["details"]:
        cer_before = r["cer_before"]
        cer_after = r["cer_after"]
        if cer_after < cer_before - 0.001:
            status = "改善"
        elif cer_after > cer_before + 0.001:
            status = "劣化"
        else:
            status = "不变"
        detail_rows.append(
            {
                "序号": len(detail_rows) + 1,
                "ID": r["id"],
                "ASR原始": r["asr"],
                "正确文本": r["correct"],
                "纠错后": r["corrected"],
                "CER_纠错前": cer_before,
                "CER_纠错后": cer_after,
                "状态": status,
                "术语命中": f"{r['term_hits']}/{r['term_total']}",
                "实体命中": f"{r['entity_hits']}/{r['entity_total']}",
                "应用层": ",".join(str(x) for x in r["layers_applied"]),
                "延迟_ms": r["latency_ms"],
            }
        )
    df_detail = pd.DataFrame(detail_rows)

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="汇总指标", index=False)
        df_layers.to_excel(writer, sheet_name="各层触发", index=False)
        df_detail.to_excel(writer, sheet_name="逐条对比", index=False)

        wb = writer.book

        # === 汇总指标 ===
        ws1 = writer.sheets["汇总指标"]
        for col_idx in range(1, df_summary.shape[1] + 1):
            ws1.column_dimensions[get_column_letter(col_idx)].width = 16
        for cell in ws1[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # === 各层触发 ===
        ws_layers = writer.sheets["各层触发"]
        ws_layers.column_dimensions["A"].width = 16
        ws_layers.column_dimensions["B"].width = 12
        ws_layers.column_dimensions["C"].width = 12
        for cell in ws_layers[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # === 逐条对比 ===
        ws2 = writer.sheets["逐条对比"]
        ws2.column_dimensions["A"].width = 8   # 序号
        ws2.column_dimensions["B"].width = 16  # ID
        ws2.column_dimensions["C"].width = 55  # ASR原始
        ws2.column_dimensions["D"].width = 55  # 正确文本
        ws2.column_dimensions["E"].width = 55  # 纠错后
        ws2.column_dimensions["F"].width = 12  # CER前
        ws2.column_dimensions["G"].width = 12  # CER后
        ws2.column_dimensions["H"].width = 10  # 状态
        ws2.column_dimensions["I"].width = 12  # 术语
        ws2.column_dimensions["J"].width = 12  # 实体
        ws2.column_dimensions["K"].width = 16  # 应用层
        ws2.column_dimensions["L"].width = 12  # 延迟

        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        text_cols = [3, 4, 5]  # C, D, E
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        for row_idx in range(2, len(df_detail) + 2):
            for col_idx in text_cols:
                ws2.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )
            status_cell = ws2.cell(row=row_idx, column=8)
            cer_after_cell = ws2.cell(row=row_idx, column=7)
            if status_cell.value == "改善":
                status_cell.fill = green_fill
                cer_after_cell.fill = green_fill
            elif status_cell.value == "劣化":
                status_cell.fill = red_fill
                cer_after_cell.fill = red_fill
            else:
                status_cell.fill = yellow_fill
                cer_after_cell.fill = yellow_fill

    print(f"[INFO] Excel 报告已保存: {excel_path}")


def main():
    args = parse_args()

    print(f"[INFO] 加载测试集: {args.testset}")
    records = load_records(args.testset, subdirs=args.subdirs)
    print(f"[INFO] 有效样本数: {len(records)}")

    if args.limit > 0:
        records = records[:args.limit]
        print(f"[INFO] 限制处理前 {args.limit} 条")

    if not records:
        print("[ERROR] 无有效样本，退出")
        sys.exit(1)

    print("[INFO] 初始化流水线...")
    pipeline = PostCorrectionPipeline()
    pipeline.warmup()

    print(f"[INFO] 开始评估 (layers={args.layers}, semantic={args.enable_semantic})...")
    t0 = time.time()
    report = evaluate(
        records,
        pipeline,
        layers=args.layers,
        enable_semantic=args.enable_semantic,
    )
    elapsed = time.time() - t0

    # 输出摘要
    summary = report["summary"]
    print("\n" + "=" * 50)
    print("评估结果摘要")
    print("=" * 50)
    print(f"总样本数: {report['meta']['total_samples']}")
    print(f"平均CER:  {summary['avg_cer_before']:.4f} -> {summary['avg_cer_after']:.4f} "
          f"(改善 {summary['cer_improvement']:.4f})")
    print(f"术语准确率: {summary['avg_term_accuracy_before']:.4f} -> {summary['avg_term_accuracy_after']:.4f}")
    print(f"实体保真率: {summary['avg_entity_fidelity_before']:.4f} -> {summary['avg_entity_fidelity_after']:.4f}")
    print(f"改善: {summary['improved_count']} | 不变: {summary['unchanged_count']} | 劣化: {summary['degraded_count']}")
    print(f"平均延迟: {summary['avg_latency_ms']:.1f}ms")
    print(f"总耗时: {elapsed:.1f}s")
    print("=" * 50)

    # 保存报告
    args.output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    report_path = args.output_dir / f"evaluation_{timestamp}.json"
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"[INFO] 报告已保存: {report_path}")

    # 导出 Excel 对比表
    excel_path = args.output_dir / f"evaluation_{timestamp}.xlsx"
    export_excel_report(report, excel_path)


if __name__ == "__main__":
    main()
