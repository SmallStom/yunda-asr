"""qwen3-asr (vLLM) + DeepFilterNet + 规则纠错 + 多语义模式 全链路评估.

参考 evaluate_three_directions.py 的输出格式，但 ASR 来源改为 qwen3-asr，
并且每条样本都带上音频文件路径。

Usage:
    python scripts/evaluate_qwen3_three_directions.py --limit 3
    python scripts/evaluate_qwen3_three_directions.py --subdirs 2
    python scripts/evaluate_qwen3_three_directions.py --modes baseline rag harness nbest
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from pathlib import Path
from typing import List, Optional

sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.audio_preprocessor import denoise_file
from src.pipeline import PostCorrectionPipeline
from tests.utils.metrics import cer


DATASET_ROOT = Path("dataset/语音识别样本集/录音普通话-长文本")
HOTWORDS_PATH = Path("data/lexicon/hotwords.json")
DEFAULT_BASE_URL = "http://192.168.1.119:8014"
DEFAULT_MODEL = "/models/Qwen3-ASR-1.7B"
MODES = ["baseline", "rag", "harness"]


def parse_args():
    parser = argparse.ArgumentParser(description="qwen3-asr 多语义模式全链路评估")
    parser.add_argument("--testset", type=Path, default=Path("data/asr_testset/asr_test_pairs_long.jsonl"))
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL)
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--subdirs", nargs="+", type=str, default=None, help="按子目录过滤 (如: 1 2 3)")
    parser.add_argument("--no-denoise", action="store_true", help="禁用 DeepFilterNet 降噪")
    parser.add_argument("--modes", nargs="+", type=str, default=MODES, help="语义精修模式")
    parser.add_argument("--output-dir", type=Path, default=Path("reports"))
    return parser.parse_args()


def load_hotwords(path: Path = HOTWORDS_PATH, max_words: int = 300) -> str:
    if not path.exists():
        return ""
    if path.suffix == ".json":
        words = json.loads(path.read_text(encoding="utf-8"))
    else:
        words = [w.strip() for w in path.read_text(encoding="utf-8").splitlines() if w.strip()]
    return ",".join(words[:max_words])


def load_records(path: Path, limit: int = 0, subdirs: Optional[List[str]] = None) -> List[dict]:
    records = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                record = json.loads(line)
            except json.JSONDecodeError:
                continue
            if subdirs:
                rid = record.get("id", "")
                if not any(rid.startswith(d + "/") for d in subdirs):
                    continue
            records.append(record)
    if limit > 0:
        records = records[:limit]
    return records


def resolve_audio(record_id: str) -> Optional[Path]:
    parts = record_id.split("/")
    if len(parts) != 2:
        return None
    folder, name = parts
    candidate = DATASET_ROOT / folder / "录音" / f"{name}.wav"
    return candidate if candidate.exists() else None


def transcribe_qwen3(audio_path: Path, hotwords: str, base_url: str, model: str, timeout: int = 120) -> str:
    url = f"{base_url}/v1/audio/transcriptions"
    headers = {"Authorization": "Bearer dummy-key-for-local"}
    data = {
        "model": model,
        "language": "zh",
        "response_format": "json",
        "temperature": 0.0,
    }
    if hotwords.strip():
        hw_list = [w.strip() for w in hotwords.replace("，", ",").split(",") if w.strip()][:100]
        data["prompt"] = "请准确转写以下铁路调度相关音频。注意识别这些术语：" + "、".join(hw_list)

    with open(audio_path, "rb") as f:
        files = {"file": (audio_path.name, f, "audio/wav")}
        response = requests.post(url, headers=headers, data=data, files=files, timeout=timeout)

    if response.status_code != 200:
        raise RuntimeError(f"qwen3-asr 请求失败: {response.status_code} {response.text[:800]}")
    return response.json().get("text", "")


def main():
    args = parse_args()

    hotwords = load_hotwords()
    print(f"[INFO] 加载热词 {len(hotwords.split(',')) if hotwords else 0} 个")

    records = load_records(args.testset, args.limit, subdirs=args.subdirs)
    print(f"[INFO] 测试样本数: {len(records)}")
    if not records:
        print("[ERROR] 无测试样本")
        sys.exit(1)

    print("[INFO] 初始化后处理流水线...")
    pipeline = PostCorrectionPipeline()
    pipeline.warmup()

    modes = args.modes
    print(f"[INFO] 开始评估 (denoise={not args.no_denoise}, modes={modes})...")

    mode_stats = {
        m: {
            "total_cer_raw": 0.0,
            "total_cer_rule": 0.0,
            "total_cer_llm": 0.0,
            "improved": 0,
            "degraded": 0,
            "unchanged": 0,
            "total_time": 0.0,
        }
        for m in modes
    }
    sample_details = []
    failed = 0

    n = len(records)
    for idx, record in enumerate(records, 1):
        rid = record.get("id", "")
        correct = record.get("correct", "")
        audio_path = resolve_audio(rid)
        if audio_path is None:
            print(f"  [{idx}/{n}] 跳过 {rid}: 找不到音频")
            failed += 1
            continue

        print(f"  [{idx}/{n}] {rid} -> {audio_path.name}")

        # 1. 可选降噪
        input_audio = audio_path
        if not args.no_denoise:
            denoised_path = Path("outputs/qwen3_eval_denoised") / f"{audio_path.stem}_att50.wav"
            denoised_path.parent.mkdir(parents=True, exist_ok=True)
            try:
                denoise_file(str(audio_path), str(denoised_path), method="deepfilternet")
                input_audio = denoised_path
            except Exception as e:
                print(f"    降噪失败，使用原音: {e}")

        # 2. qwen3-asr
        try:
            raw_text = transcribe_qwen3(input_audio, hotwords, args.base_url, args.model)
        except Exception as e:
            print(f"    ASR失败: {e}")
            failed += 1
            continue

        cer_raw = cer(raw_text, correct)

        # 3. 规则纠错
        rule_result = pipeline.run(raw_text, layers=[1, 2, 3], enable_semantic=False)
        rule_text = rule_result.corrected
        cer_rule = cer(rule_text, correct)

        detail = {
            "id": rid,
            "audio": str(audio_path),
            "correct": correct,
            "asr_raw": raw_text,
            "rule_corrected": rule_text,
            "cer_raw": round(cer_raw, 4),
            "cer_rule": round(cer_rule, 4),
            "modes": {},
        }

        # 4. 各语义模式
        for mode in modes:
            t0 = time.time()
            try:
                llm_result = pipeline.run(
                    raw_text,
                    layers=[1, 2, 3],
                    enable_semantic=True,
                    semantic_mode=mode,
                )
                llm_text = llm_result.corrected
                cer_llm = cer(llm_text, correct)
            except Exception as e:
                print(f"    [错误] {mode}: {type(e).__name__}: {e}")
                llm_text = rule_text
                cer_llm = cer_rule
            llm_time = time.time() - t0

            mode_stats[mode]["total_cer_raw"] += cer_raw
            mode_stats[mode]["total_cer_rule"] += cer_rule
            mode_stats[mode]["total_cer_llm"] += cer_llm
            mode_stats[mode]["total_time"] += llm_time

            if cer_llm < cer_rule - 0.001:
                mode_stats[mode]["improved"] += 1
                status = "改善"
            elif cer_llm > cer_rule + 0.001:
                mode_stats[mode]["degraded"] += 1
                status = "劣化"
            else:
                mode_stats[mode]["unchanged"] += 1
                status = "不变"

            detail["modes"][mode] = {
                "corrected": llm_text,
                "cer": round(cer_llm, 4),
                "status": status,
                "time_ms": round(llm_time * 1000, 1),
            }

        sample_details.append(detail)

    valid_n = len(sample_details)
    if valid_n == 0:
        print("[ERROR] 无成功样本")
        sys.exit(1)

    # 汇总打印
    print(f"\n{'='*100}")
    print("qwen3-asr 多语义模式全链路汇总")
    print(f"{'='*100}")
    print(f"{'模式':<12s} {'原始CER':>8s} {'规则CER':>8s} {'LLM CER':>8s} "
          f"{'LLM改善':>8s} {'改善':>6s} {'劣化':>6s} {'不变':>6s} {'平均耗时':>10s}")
    print("-" * 100)
    for mode in modes:
        s = mode_stats[mode]
        avg_raw = s["total_cer_raw"] / valid_n
        avg_rule = s["total_cer_rule"] / valid_n
        avg_llm = s["total_cer_llm"] / valid_n
        avg_time = s["total_time"] / valid_n * 1000
        print(f"{mode:<12s} {avg_raw:>8.4f} {avg_rule:>8.4f} {avg_llm:>8.4f} "
              f"{avg_rule - avg_llm:>+8.4f} {s['improved']:>6d} {s['degraded']:>6d} "
              f"{s['unchanged']:>6d} {avg_time:>8.0f}ms")

    # 写 Excel
    args.output_dir.mkdir(parents=True, exist_ok=True)
    ts = int(time.time())

    # Sheet1
    summary_rows = []
    for mode in modes:
        s = mode_stats[mode]
        summary_rows.append({
            "模式": mode,
            "原始CER": round(s["total_cer_raw"] / valid_n, 4),
            "规则CER": round(s["total_cer_rule"] / valid_n, 4),
            "LLM CER": round(s["total_cer_llm"] / valid_n, 4),
            "LLM改善": round(s["total_cer_rule"] / valid_n - s["total_cer_llm"] / valid_n, 4),
            "改善数": s["improved"],
            "劣化数": s["degraded"],
            "不变数": s["unchanged"],
            "平均耗时(ms)": round(s["total_time"] / valid_n * 1000, 1),
        })
    df_summary = pd.DataFrame(summary_rows)

    # Sheet2
    detail_rows = []
    for i, detail in enumerate(sample_details, 1):
        row = {
            "序号": i,
            "ID": detail["id"],
            "音频路径": detail["audio"],
            "正确文本": detail["correct"],
            "qwen3-asr原始": detail["asr_raw"],
            "原始CER": detail["cer_raw"],
            "规则纠错": detail["rule_corrected"],
            "规则CER": detail["cer_rule"],
        }
        for mode in modes:
            m = detail["modes"][mode]
            row[f"{mode}_纠错结果"] = m["corrected"]
            row[f"{mode}_CER"] = m["cer"]
            row[f"{mode}_状态"] = m["status"]
        detail_rows.append(row)
    df_detail = pd.DataFrame(detail_rows)

    xlsx_path = args.output_dir / f"qwen3_three_directions_{ts}.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="汇总指标", index=False)
        df_detail.to_excel(writer, sheet_name="逐条对比", index=False)

        wb = writer.book

        # 汇总指标格式
        ws1 = writer.sheets["汇总指标"]
        ws1.column_dimensions["A"].width = 12
        for col_idx in range(2, 10):
            ws1.column_dimensions[get_column_letter(col_idx)].width = 12
        for cell in ws1[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # 逐条对比格式
        ws2 = writer.sheets["逐条对比"]
        ws2.column_dimensions["A"].width = 6   # 序号
        ws2.column_dimensions["B"].width = 16  # ID
        ws2.column_dimensions["C"].width = 45  # 音频路径
        ws2.column_dimensions["D"].width = 50  # 正确文本
        ws2.column_dimensions["E"].width = 50  # qwen3-asr原始
        ws2.column_dimensions["F"].width = 10  # 原始CER
        ws2.column_dimensions["G"].width = 50  # 规则纠错
        ws2.column_dimensions["H"].width = 10  # 规则CER

        col_offset = 9  # 从 I 列开始是各模式数据
        for j, mode in enumerate(modes):
            base = col_offset + j * 3
            ws2.column_dimensions[get_column_letter(base)].width = 50      # 纠错结果
            ws2.column_dimensions[get_column_letter(base + 1)].width = 10  # CER
            ws2.column_dimensions[get_column_letter(base + 2)].width = 8   # 状态

        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        text_cols = [3, 4, 5, 7]  # C, D, E, G
        for j, mode in enumerate(modes):
            text_cols.append(col_offset + j * 3)
        for row_idx in range(2, len(df_detail) + 2):
            for col_idx in text_cols:
                ws2.cell(row=row_idx, column=col_idx).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        for row_idx in range(2, len(df_detail) + 2):
            for j, mode in enumerate(modes):
                status_col = col_offset + j * 3 + 2
                cer_col = col_offset + j * 3 + 1
                cell = ws2.cell(row=row_idx, column=status_col)
                cer_cell = ws2.cell(row=row_idx, column=cer_col)
                if cell.value == "改善":
                    cell.fill = green_fill
                    cer_cell.fill = green_fill
                elif cell.value == "劣化":
                    cell.fill = red_fill
                    cer_cell.fill = red_fill
                else:
                    cell.fill = yellow_fill
                    cer_cell.fill = yellow_fill

    print(f"\n[INFO] Excel 报告已保存: {xlsx_path}")


if __name__ == "__main__":
    main()
